# Databricks notebook source
if spark.version.replace('.','') < '31':
  dbutils.library.installPyPI("openpyxl")
else:
  %pip install openpyxl

# COMMAND ----------

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
import os
import shutil
import pandas as pd

# COMMAND ----------

def OpenExcelFromLake (path_lake):

  # define as variáveis utilizadas a partir do argumento informado
  file = path_lake.split('/')[-1]
  file_name = file.split('.')[0]
  file_ext = file.split('.')[1]
  file_name_tmp = str(time.time()).replace('.','-') + '.' + file_ext
  path_dbfs ="/dbfs/tmp/" + file_name_tmp

  # copia o xlsx para o dbfs
  dbutils.fs.cp(path_lake, path_dbfs)
  
  # cria um objeto workbook a partir do arquivo excel no dbfs
  wb = load_workbook('/dbfs' + path_dbfs)
  
  return wb

# COMMAND ----------

def SaveExcelToLake (ObjectWB, path_lake):

  resultado = ""

  try:
    # determina as variáveis
    nome = str(time.time()).replace('.','-') + '.xlsx'
    file_name_local_tmp = '/local_disk0/tmp/' + nome
    file_name_dbfs_tmp = '/dbfs/tmp/' + nome

    # salva o objeto no drive local
    wb.save(file_name_local_tmp)   
    
    # copia o arquivo excel do disco local para o dbfs
    shutil.copyfile(file_name_local_tmp, file_name_dbfs_tmp)  

    # copia o arquivo do dbfs para o data lake
    dbutils.fs.cp(file_name_dbfs_tmp.replace('/dbfs',''), path_lake)

    # deleta o arquivo do dbfs
    dbutils.fs.rm(file_name_dbfs_tmp)

    # deleta o arquivo do disco local
    os.remove(file_name_local_tmp)

    # mensagem de saída de sucesso
    resultado = "Arquivo salvo: " + path_lake    
            
  except Exception as e:
    print(e)
  
  return resultado

# COMMAND ----------

def OpenPandasFromLakeExcel (path_lake):

  # define as variáveis utilizadas a partir do argumento informado
  file = path_lake.split('/')[-1]
  file_name = file.split('.')[0]
  file_ext = file.split('.')[1]
  file_name_tmp = str(time.time()).replace('.','-') + '.' + file_ext
  path_dbfs ="/dbfs/tmp/" + file_name_tmp

  # copia o xlsx para o dbfs
  dbutils.fs.cp(path_lake, path_dbfs)
  
  # abre o arquivo e devolve um dataframe pandas
  df = pd.read_excel('/dbfs' + path_dbfs, sheet_name='Sheet')
  
  return df
